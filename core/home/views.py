import math
import pandas as pd
from django.db import connection
from .models import Client, TenderResults, ScraperControl, Search
from .serializers import ClientSerializer, TenderResultsSerializer, SearchTenderSerializer,SearchSerializer
from .tasks import run_scraper
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated,AllowAny
from django.contrib.auth.models import User


@api_view(["POST"])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")
    print("Username:", username, "Password:", password)
    user = authenticate(username=username, password=password)
    if not user:
        return Response({"error": "Invalid credentials"}, status=401)

    token, _ = Token.objects.get_or_create(user=user)
    return Response({
        "token": token.key,
        "username": user.username
    })

@api_view(["POST"])
@permission_classes([AllowAny])
def register(request):
    username = request.data.get("username")
    password = request.data.get("password")

    if not username or not password:
        return Response(
            {"error": "Username and password are required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if User.objects.filter(username=username).exists():
        return Response(
            {"error": "Username already exists"},
            status=status.HTTP_409_CONFLICT
        )

    user = User.objects.create_user(
        username=username,
        password=password
    )

    return Response(
        {
            "message": "User registered successfully",
            "user_id": user.id,
            "username": user.username
        },
        status=status.HTTP_201_CREATED
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_clients(request):
    uploaded_file = request.FILES.get("file")

    if not uploaded_file:
        return Response({"error": "No file uploaded"}, status=400)

    filename = uploaded_file.name.lower()
    if filename.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        df = pd.read_csv(uploaded_file, encoding="latin1")

    objs = [
        Client(
            state_name=row["state_name"],
            site_url=row["site_url"],
            search_key=row["search_key"],
            exclude_key=row["exclude_key"],
            user=request.user,  # For authentication
        )
        for _, row in df.iterrows()
    ]
    if connection.connection is None:
        connection.connect()

    Client.objects.bulk_create(objs, ignore_conflicts=True)
    return Response({"message": "Clients uploaded successfully"})



from django.db import transaction
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_search_tender_req(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"error": "No file uploaded"}, status=400)

    if uploaded_file.name.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        df = pd.read_csv(uploaded_file, encoding="utf-8", errors="ignore")

    df = df.fillna("")
    print("Columns:", df.columns.tolist())

    objs = []
    for _, row in df.iterrows():
        objs.append(Search(
            state_name=str(row["state_name"])[:255],
            site_url=str(row["site_url"])[:500],
            search_key=str(row["search_key"]),
            exclude_key=str(row["exclude_key"]),
            user=request.user,
        ))

    with transaction.atomic():
        created = Search.objects.bulk_create(objs)

    return Response({
        "message": "Tender search req uploaded successfully",
        "rows_saved": len(created)
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_clients(request):
    clients = Client.objects.filter(user=request.user)
    serializer = ClientSerializer(clients, many=True)
    return Response(serializer.data)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_client_search(request):
    search_clients = Search.objects.filter(user=request.user)
    serializer = SearchTenderSerializer(search_clients, many=True)
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_client_fields_excel(request):
    # Get field names excluding 'id'
    fields = [
        field.name
        for field in Search._meta.fields
        if field.name != "id"
    ]
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Fields"
    # Add header row
    ws.append(fields)
    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="client_fields.xlsx"'
    wb.save(response)
    return response



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_client_fields(request):
    queryset = Search.objects.values(
        "state_name", "site_url", "search_key", "exclude_key"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Data"

    # ✅ Header row
    headers = ["state_name", "site_url", "search_key", "exclude_key"]
    ws.append(headers)

    # ✅ Data rows
    for row in queryset:
        ws.append([
            row["state_name"],
            row["site_url"],
            row["search_key"],
            row["exclude_key"],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="search_data.xlsx"'

    wb.save(response)
    return response

# from threading import Thread
#
# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# def run_scraper_task(request, search_id):
#     user_id = request.user.id
#     print("Search_id:", search_id)
#
#     # Start scraper in background thread
#     t = Thread(
#         target=run_scraper,
#         args=(user_id, search_id),
#         daemon=True  # stops thread if server stops
#     )
#     t.start()
#
#     return Response({
#         "status": "started",
#         "message": "Scraper started in background"
#     })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def run_scraper_task(request, search_id):
    print("Search_id:", search_id)
    user_id=request.user.id
    # run_scraper(user_id, search_id)

    # if control.is_running:
    #     return Response({"message": "Scraper already running"}, status=400)

    task = run_scraper.apply_async(args=[request.user.id, search_id])
    #
    return Response({
        "message": "Scraper started",
        "task_id": task.id
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def del_scraper_task(request, search_id=None):
    try:
        user = request.user
        Search.objects.filter(pk=search_id, user=user).delete()
        return Response({"message": "Deleted Scraper Task", "task_id": search_id})
    except:
        return Response({"error": "Delete failed"})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_search_req(request):
    print("Adding Search Req:",request.data)

    serializer = SearchSerializer(
        data=request.data
    )
    if serializer.is_valid():
        serializer.save(user=request.user)  # attach logged-in user
        return Response(
            {"message": "Data Added Successfully"},
            status=status.HTTP_201_CREATED
        )
    return Response(
        serializer.errors,
        status=status.HTTP_400_BAD_REQUEST
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def scraper_status(request):
    try:
        control = ScraperControl.objects.get(user=request.user)
    except ScraperControl.DoesNotExist:
        return Response({
            "is_running": False,
            "searching_state_name": "",
            "searching_key": "",
            "task_id": None
        })

    return Response({
        "is_running": control.is_running,
        "searching_state_name": control.searching_state_name,
        "searching_key": control.searching_key,

    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def stop_scraper_task(request):
    control = ScraperControl.objects.filter(user=request.user).first()
    # if control:
    #     control.is_running = False
    #     control.save()
    #     return Response({"message": "Scraper stopping..."})

    return Response({"message": "No scraper running"}, status=400)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_tenders(request):
    client_reqs = Search.objects.filter(user=request.user).values(
        "state_name", "search_key", "exclude_key"
    )

    tender_filter = Q()
    final_exclude = Q()

    for req in client_reqs:
        state = req["state_name"]

        search_keys = [k.strip() for k in req["search_key"].split(",") if k.strip()]
        exclude_keys = [e.strip() for e in req["exclude_key"].split("|") if e.strip()]

        search_q = Q()
        for key in search_keys:
            search_q |= Q(search_key__iexact=key)

        tender_filter |= Q(state_name__iexact=state) & search_q

        exclude_q = Q()
        for ex in exclude_keys:
            exclude_q |= Q(work_description__icontains=ex)

        final_exclude |= Q(state_name__iexact=state) & exclude_q

    tenders = TenderResults.objects.filter(tender_filter).exclude(final_exclude).distinct()

    # 🔍 UI FILTERS
    if request.GET.get("tender_id"):
        tenders = tenders.filter(tender_id__icontains=request.GET["tender_id"])

    if request.GET.get("state_name"):
        tenders = tenders.filter(state_name__iexact=request.GET["state_name"])

    if request.GET.get("organization"):
        tenders = tenders.filter(organization_chain__icontains=request.GET["organization"])

    paginator = TenderPagination()
    page = paginator.paginate_queryset(
        tenders.order_by("-search_time"), request
    )
    if client_reqs:
        serializer = TenderResultsSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)



class TenderPagination(PageNumberPagination):
    page_size = 10

    def get_paginated_response(self, data):
        total_pages = math.ceil(self.page.paginator.count / self.page_size)

        return Response({
            "count": self.page.paginator.count,
            "total_pages": total_pages,
            "current_page": self.page.number,
            "next": self.get_next_link(),
            "previous": self.get_previous_link(),
            "results": data,
        })



@api_view(["POST"])
@permission_classes([IsAuthenticated])
def del_req(request):
    Client.objects.filter(user=request.user).delete()
    return Response({"message": " All Requirements  deleted"})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def del_search(request):
    Search.objects.filter(user=request.user).delete()
    return Response({"message": " All Search Req  deleted"})



from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_all_tenders(request):
    client_reqs = Client.objects.filter(user=request.user).values(
        "state_name", "search_key", "exclude_key"
    )

    final_filter = Q()
    final_exclude = Q()

    for client_req in client_reqs:
        state = client_req["state_name"]

        search_keys = [
            k.strip() for k in client_req["search_key"].split(",") if k.strip()
        ]

        exclude_keys = [
            e.strip() for e in client_req["exclude_key"].split("|") if e.strip()
        ]

        # INCLUDE keys
        search_q = Q()
        for key in search_keys:
            search_q |= Q(search_key__iexact=key)

        final_filter |= Q(state_name__iexact=state) & search_q

        # EXCLUDE keys
        exclude_q = Q()
        for x_key in exclude_keys:
            exclude_q |= Q(work_description__icontains=x_key)
        final_exclude |= Q(state_name__iexact=state) & exclude_q

    tenders = (
        TenderResults.objects
        .filter(final_filter)
        .exclude(final_exclude)
        .distinct()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    headers = [
        "search_time", "tender_id", "state_name", "search_key",
        "site_link", "work_description", "organization_chain",
        "bid_submission_end_date", "bid_submission_end_time",
        "tender_value", "emd_amt", "tender_fee"
    ]
    ws.append(headers)
    for t in tenders:
        ws.append([
            t.search_time,
            t.tender_id,
            t.state_name,
            t.search_key,
            t.site_link,
            t.work_description,
            t.organization_chain,
            t.bid_submission_end_date,
            t.bid_submission_end_time,
            t.tender_value,
            t.emd_amt,
            t.tender_fee,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="all_tenders.xlsx"'
    wb.save(response)
    return response






# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def get_tenders(request):
#     print("User:",request.user)
#     client_reqs = Client.objects.filter(user=request.user).values(
#         "state_name", "search_key", "exclude_key"
#     )
#
#     tender_filter = Q()
#     for client_req in client_reqs:
#         state = client_req["state_name"]
#         search_keys = [k.strip() for k in client_req["search_key"].split(",") if k.strip()]
#
#         # Build search_key OR conditions
#         search_q = Q()
#         for key in search_keys:
#             search_q |= Q(search_key__iexact=key)
#         # Combine state + search keys
#         tender_filter |= (Q(state_name__iexact=state) & search_q)
#     # print("tender_filter:",tender_filter)
#     tenders = TenderResults.objects.filter(tender_filter).distinct()
#     # print("Tenders:",tenders)
#     for tender in tenders:
#         print("Tender Search Time:",tender.search_time)
#         print("Tender ID:",tender.tender_id)
#         print("State Name:",tender.state_name)
#
#     serializer = TenderResultsSerializer(tenders, many=True)
#     return Response(serializer.data)



# @api_view(["GET"])  # Pagination
# @permission_classes([IsAuthenticated])
# def get_tenders(request):#Previously  Client.objects.filter
#     client_reqs = Search.objects.filter(user=request.user).values(
#         "state_name", "search_key", "exclude_key"
#     )
#     final_exclude= Q()
#     tender_filter = Q()
#     for client_req in client_reqs:
#         state = client_req["state_name"]
#         search_keys = [
#             k.strip() for k in client_req["search_key"].split(",") if k.strip()
#         ]
#         exclude_keys = [
#             e.strip() for e in client_req["exclude_key"].split("|") if e.strip()
#         ]
#         search_q = Q()
#         for key in search_keys:
#             search_q |= Q(search_key__iexact=key)
#         tender_filter |= (Q(state_name__iexact=state) & search_q)
#
#         exclude_q = Q()
#         for x_key in exclude_keys:
#             exclude_q |= Q(work_description__icontains=x_key)
#         final_exclude |= Q(state_name__iexact=state) & exclude_q
#
#     if client_reqs:
#         tenders = TenderResults.objects.filter(tender_filter).exclude(final_exclude).distinct().order_by("-search_time")
#         paginator = TenderPagination()
#         page = paginator.paginate_queryset(tenders, request)
#
#         serializer = TenderResultsSerializer(page, many=True)
#         return paginator.get_paginated_response(serializer.data)



# @api_view(["POST"])
# @permission_classes([IsAuthenticated])
# def del_tenders(request):
#     TenderResults.objects.all().delete()
#     return Response({"message": " All Tender deleted"})





#
# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def download_all_tenders(request):
#     client_reqs = Client.objects.filter(user=request.user).values(
#         "state_name", "search_key"
#     )
#
#     tender_filter = Q()
#     for client_req in client_reqs:
#         state = client_req["state_name"]
#         search_keys = [
#             k.strip() for k in client_req["search_key"].split(",") if k.strip()
#         ]
#
#         search_q = Q()
#         for key in search_keys:
#             search_q |= Q(search_key__iexact=key)
#
#         tender_filter |= (Q(state_name__iexact=state) & search_q)
#
#     tenders = TenderResults.objects.filter(tender_filter).distinct()
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Tenders"
#
#     headers = [
#         "search_time", "tender_id", "state_name", "search_key",
#         "site_link", "work_description", "organization_chain",
#         "bid_submission_end_date", "bid_submission_end_time",
#         "tender_value", "emd_amt", "tender_fee"
#     ]
#     ws.append(headers)
#
#     for t in tenders:
#         ws.append([
#             t.search_time,
#             t.tender_id,
#             t.state_name,
#             t.search_key,
#             t.site_link,
#             t.work_description,
#             t.organization_chain,
#             t.bid_submission_end_date,
#             t.bid_submission_end_time,
#             t.tender_value,
#             t.emd_amt,
#             t.tender_fee,
#         ])
#
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="all_tenders.xlsx"'
#
#     wb.save(response)
#     return response
